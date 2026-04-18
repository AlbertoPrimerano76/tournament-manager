from __future__ import annotations

import re
from io import BytesIO
from zoneinfo import ZoneInfo

from app.schemas.program import (
    AgeGroupProgramResponse,
    ProgramGroupResponse,
    ProgramMatchResponse,
    ProgramPhaseResponse,
)

# ── Column layout (A=1 … O=15 in openpyxl 1-based, 0-based indices below) ─────
#  A    B       C      D          E          F  G    H          I          J  K  L  M       N       O
#  ORA  CAMPO   [img]  COD_HOME   NOME_HOME  -  vs   COD_AWAY   NOME_AWAY  -  -  -  PT_H    PT_A    ARBITRO
#
# Column C (index 2) is a narrow image-lane used for team-crest thumbnails in
# the team-roster block.  It stays empty in match rows.
_NCOLS = 15
_COL_ORA       = 0   # A
_COL_CAMPO     = 1   # B
_COL_IMG       = 2   # C  ← team crest lane
_COL_COD_HOME  = 3   # D
_COL_NOM_HOME  = 4   # E
_COL_VS        = 6   # G
_COL_COD_AWAY  = 7   # H
_COL_NOM_AWAY  = 8   # I
_COL_PT_H      = 12  # M  punti casa
_COL_PT_A      = 13  # N  punti ospite
_COL_ARBITRO   = 14  # O

# Header occupies rows 1-4 (logo row, subtitle, group name, spacer/border)
_HEADER_ROWS = 3


def _safe_filename(value: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower())
    return normalized.strip("-") or "categoria"


def _safe_sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "", value).strip()
    return cleaned[:31] or "Foglio"


def _format_time(value, tournament_timezone: str, fallback: str = "") -> str:
    if value is None:
        return fallback
    try:
        tz = ZoneInfo(tournament_timezone)
    except Exception:
        tz = ZoneInfo("Europe/Rome")
    if getattr(value, "tzinfo", None) is None:
        return value.strftime("%H:%M")
    return value.astimezone(tz).strftime("%H:%M")


def _field_label(match: ProgramMatchResponse) -> str:
    if not match.field_name:
        return ""
    parts = [match.field_name]
    if match.field_number is not None:
        parts.append(f"Campo {match.field_number}")
    return " · ".join(parts)


def _sort_matches(matches: list[ProgramMatchResponse]) -> list[ProgramMatchResponse]:
    return sorted(
        matches,
        key=lambda m: (
            m.scheduled_at.isoformat() if m.scheduled_at else "9999-12-31T23:59:59",
            m.field_name or "",
            m.field_number or 0,
            m.bracket_position or 0,
        ),
    )


def _round_duration_suffix(
    round_matches: list[ProgramMatchResponse],
    default_duration_minutes: int | None = None,
) -> str:
    durations = {
        match.match_duration_minutes
        for match in round_matches
        if match.match_duration_minutes is not None
    }
    if len(durations) != 1:
        return ""
    duration = next(iter(durations))
    if default_duration_minutes is not None and duration == default_duration_minutes:
        return ""
    return f"  ·  {duration} min/partita"


# ── Image helpers ──────────────────────────────────────────────────────────────

def _fetch_image_bytes(url: str | None, cache: dict) -> bytes | None:
    """Download image bytes from URL, with a simple dict-based cache."""
    if not url:
        return None
    if url in cache:
        return cache[url]
    try:
        from urllib.request import Request, urlopen
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=6) as resp:
            data: bytes | None = resp.read()
    except Exception:
        data = None
    cache[url] = data
    return data


def _to_png_bytes(img_bytes: bytes) -> bytes:
    """Convert image bytes to PNG (openpyxl only supports JPEG/PNG/GIF/BMP/TIFF)."""
    try:
        from PIL import Image as PILImage
        with PILImage.open(BytesIO(img_bytes)) as im:
            buf = BytesIO()
            im.convert("RGBA").save(buf, format="PNG")
            return buf.getvalue()
    except Exception:
        return img_bytes


def _add_image(ws, img_bytes: bytes | None, cell_ref: str, w_px: int, h_px: int) -> None:
    """Anchor an image to *cell_ref* at *w_px × h_px* pixels (no-op on error)."""
    if not img_bytes:
        return
    try:
        from openpyxl.drawing.image import Image as XlImage
        img = XlImage(BytesIO(_to_png_bytes(img_bytes)))
        img.width = w_px
        img.height = h_px
        img.anchor = cell_ref
        ws.add_image(img)
    except Exception:
        pass


# ── Styling helpers ────────────────────────────────────────────────────────────

def _set_col_widths(ws) -> None:
    widths = {
        "A": 8,    # ORA / code
        "B": 30,   # CAMPO / team name
        "C": 4,    # image lane (crest ~20 px ≈ width 3-4)
        "D": 5,    # code home
        "E": 26,   # nome home
        "F": 2,    # spacer
        "G": 4,    # vs
        "H": 5,    # code away
        "I": 26,   # nome away
        "J": 2,
        "K": 2,
        "L": 2,
        "M": 9,    # mete home
        "N": 9,    # mete away
        "O": 22,   # ARBITRO
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _write_sheet_header(
    ws,
    tournament_name: str,
    age_group_label: str,
    phase_info: str,
    org_logo: bytes | None,
    tournament_logo: bytes | None,
) -> None:
    """Write the 3-row decorative header (logo · title · subtitle)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill

    # ── Row 1: org-logo | tournament name | tournament-logo ───────────────────
    ws.row_dimensions[1].height = 52

    # Background for entire header row
    navy_fill = PatternFill("solid", fgColor="1E3A5F")
    thin = Side(style="thin", color="2D5A8E")
    for col in range(1, _NCOLS + 1):
        c = ws.cell(row=1, column=col)
        c.fill = navy_fill
        c.border = Border(bottom=thin)

    # Tournament name — merged columns 3-13 (C1:M1)
    title_cell = ws.cell(row=1, column=3, value=tournament_name.upper())
    title_cell.font = Font(bold=True, size=20, color="FFFFFF", name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=13)

    # Logos: place images so they sit nicely inside the tall row
    _add_image(ws, org_logo,         "A1", 42, 42)
    _add_image(ws, tournament_logo,  "N1", 42, 42)

    # ── Row 2: age-group · phase subtitle ─────────────────────────────────────
    ws.row_dimensions[2].height = 22
    light_navy = PatternFill("solid", fgColor="2D4F7C")
    for col in range(1, _NCOLS + 1):
        c = ws.cell(row=2, column=col)
        c.fill = light_navy
    sub_cell = ws.cell(row=2, column=1, value=f"{age_group_label}   ·   {phase_info}")
    sub_cell.font = Font(bold=True, size=11, color="D1E4F6", name="Calibri")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_NCOLS)

    # ── Row 3: thin accent bar ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 5
    accent = PatternFill("solid", fgColor="F59E0B")
    for col in range(1, _NCOLS + 1):
        ws.cell(row=3, column=col).fill = accent


def _style_section_header(ws, row: int, hex_bg: str = "DBEAFE") -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="93C5FD")
    border = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color="3B82F6"))
    fill = PatternFill("solid", fgColor=hex_bg)
    for col in range(1, _NCOLS + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, size=9, color="1E3A5F", name="Calibri")
        c.fill = fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _style_group_name_row(ws, row: int) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fill = PatternFill("solid", fgColor="1D4ED8")
    thin = Side(style="thin", color="3B82F6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, _NCOLS + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.border = border
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")


def _style_match_row(ws, row: int, alt: bool, row_height: int = 18) -> None:
    from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="F0F4FF" if alt else "FFFFFF")
    ws.row_dimensions[row].height = row_height
    for col in range(1, _NCOLS + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.border = border
        c.alignment = Alignment(vertical="center")
        c.font = Font(size=9, name="Calibri")
    for col in [_COL_ORA + 1, _COL_CAMPO + 1, _COL_IMG + 1, _COL_COD_HOME + 1, _COL_VS + 1,
                _COL_COD_AWAY + 1, _COL_PT_H + 1, _COL_PT_A + 1]:
        c = ws.cell(row=row, column=col)
        c.alignment = Alignment(horizontal="center", vertical="center")
    # Bold team codes
    for col in [_COL_COD_HOME + 1, _COL_COD_AWAY + 1]:
        ws.cell(row=row, column=col).font = Font(bold=True, size=9, color="1E3A5F", name="Calibri")
    # Score cells — thick border, extra padding implied by row height
    from openpyxl.styles import Border as Brd, Side as Sd
    thick = Sd(style="medium", color="94A3B8")
    for col in [_COL_PT_H + 1, _COL_PT_A + 1]:
        ws.cell(row=row, column=col).border = Brd(left=thick, right=thick, top=thick, bottom=thick)


# ── Group sheet ────────────────────────────────────────────────────────────────

def _write_group_sheet(
    ws,
    tournament_name: str,
    age_group_label: str,
    group: ProgramGroupResponse,
    group_num: int,
    phase_name: str,
    tournament_timezone: str,
    org_logo: bytes | None,
    tournament_logo: bytes | None,
    image_cache: dict,
    match_duration_minutes: int | None = None,
) -> None:
    from openpyxl.styles import Font, Alignment

    # Build team-code map: tournament_team_id → (code, label)
    team_code: dict[str, tuple[str, str]] = {}
    for pos, team in enumerate(group.teams):
        code = f"{chr(65 + pos)}{group_num}"
        if team.tournament_team_id:
            team_code[team.tournament_team_id] = (code, team.label)

    _set_col_widths(ws)
    ws.freeze_panes = f"A{_HEADER_ROWS + 5}"  # freeze past header + team list header

    # ── Header block (rows 1-3) ────────────────────────────────────────────────
    _write_sheet_header(ws, tournament_name, age_group_label, phase_name, org_logo, tournament_logo)

    # ── Row 4: group name banner ───────────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    banner_text = group.name
    if match_duration_minutes:
        banner_text = f"{group.name}  ·  {match_duration_minutes} min/partita"
    ws.cell(row=4, column=1, value=banner_text)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=_NCOLS)
    _style_group_name_row(ws, 4)

    # ── Row 5: team-roster column headers ─────────────────────────────────────
    teams_header = [""] * _NCOLS
    teams_header[0] = "COD."
    teams_header[2] = ""     # image lane header (blank)
    teams_header[1] = "SQUADRA"
    ws.append(teams_header)
    _style_section_header(ws, 5, "DBEAFE")

    # ── Rows 6+: team list with crests ────────────────────────────────────────
    for pos, team in enumerate(group.teams):
        code = f"{chr(65 + pos)}{group_num}"
        row_data = [""] * _NCOLS
        row_data[0] = code
        row_data[1] = team.label
        ws.append(row_data)
        current_row = ws.max_row
        _style_match_row(ws, current_row, pos % 2 == 1, row_height=26)

        # Insert team crest in the image lane (col C = index 2)
        logo_bytes = _fetch_image_bytes(getattr(team, "team_logo_url", None), image_cache)
        _add_image(ws, logo_bytes, f"C{current_row}", 20, 20)

    # ── Spacer row ─────────────────────────────────────────────────────────────
    ws.append([""] * _NCOLS)
    ws.row_dimensions[ws.max_row].height = 8

    # ── Match schedule column headers ──────────────────────────────────────────
    header = [""] * _NCOLS
    header[_COL_ORA]      = "ORA"
    header[_COL_CAMPO]    = "CAMPO"
    header[_COL_COD_HOME] = ""
    header[_COL_NOM_HOME] = "SQUADRA CASA"
    header[_COL_VS]       = "vs"
    header[_COL_COD_AWAY] = ""
    header[_COL_NOM_AWAY] = "SQUADRA OSPITE"
    header[_COL_PT_H]   = "PT"
    header[_COL_PT_A]   = "PT"
    header[_COL_ARBITRO]  = "ARBITRO"
    ws.append(header)
    _style_section_header(ws, ws.max_row, "DBEAFE")

    # ── Match rows ─────────────────────────────────────────────────────────────
    matches = _sort_matches(group.matches)
    for i, match in enumerate(matches):
        h_code, h_name = team_code.get(match.home_team_id or "", ("", match.home_label or ""))
        a_code, a_name = team_code.get(match.away_team_id or "", ("", match.away_label or ""))
        row_data = [""] * _NCOLS
        row_data[_COL_ORA]       = _format_time(match.scheduled_at, tournament_timezone)
        row_data[_COL_CAMPO]     = _field_label(match)
        row_data[_COL_COD_HOME]  = h_code
        row_data[_COL_NOM_HOME]  = h_name
        row_data[_COL_VS]        = "vs"
        row_data[_COL_COD_AWAY]  = a_code
        row_data[_COL_NOM_AWAY]  = a_name
        row_data[_COL_PT_H]    = match.home_score if match.home_score is not None else ""
        row_data[_COL_PT_A]    = match.away_score if match.away_score is not None else ""
        row_data[_COL_ARBITRO]   = match.referee or ""
        ws.append(row_data)
        _style_match_row(ws, ws.max_row, i % 2 == 1)

        # Small home/away crests next to team names in match rows
        if match.home_team_id and match.home_logo_url:
            logo = _fetch_image_bytes(match.home_logo_url, image_cache)
            _add_image(ws, logo, f"C{ws.max_row}", 14, 14)


# ── Knockout sheet ─────────────────────────────────────────────────────────────

def _write_knockout_sheet(
    ws,
    tournament_name: str,
    age_group_label: str,
    phase: ProgramPhaseResponse,
    tournament_timezone: str,
    org_logo: bytes | None,
    tournament_logo: bytes | None,
) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _set_col_widths(ws)
    ws.freeze_panes = f"A{_HEADER_ROWS + 1}"

    _write_sheet_header(ws, tournament_name, age_group_label, phase.name, org_logo, tournament_logo)

    matches = _sort_matches(phase.knockout_matches)
    matches_by_round: dict[str, list[ProgramMatchResponse]] = {}
    for m in matches:
        rname = m.bracket_round or phase.name
        matches_by_round.setdefault(rname, []).append(m)

    current_row = _HEADER_ROWS + 1

    for round_name, round_matches in matches_by_round.items():
        # Round sub-header — amber band
        ws.row_dimensions[current_row].height = 18
        amber_fill = PatternFill("solid", fgColor="92400E")
        thin = Side(style="thin", color="B45309")
        for col in range(1, _NCOLS + 1):
            c = ws.cell(row=current_row, column=col)
            c.fill = amber_fill
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        rnd_cell = ws.cell(
            row=current_row,
            column=1,
            value=f"{round_name.upper()}{_round_duration_suffix(round_matches, phase.match_duration_minutes)}",
        )
        rnd_cell.font = Font(bold=True, size=10, color="FEF3C7", name="Calibri")
        rnd_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=_NCOLS)
        current_row += 1

        # Column header
        header = [""] * _NCOLS
        header[_COL_ORA]      = "ORA"
        header[_COL_CAMPO]    = "CAMPO"
        header[_COL_NOM_HOME] = "SQUADRA CASA"
        header[_COL_VS]       = "vs"
        header[_COL_NOM_AWAY] = "SQUADRA OSPITE"
        header[_COL_PT_H]   = "PT"
        header[_COL_PT_A]   = "PT"
        header[_COL_ARBITRO]  = "ARBITRO"
        ws.cell(row=current_row, column=1)  # ensure row exists before styling
        for ci, val in enumerate(header, start=1):
            ws.cell(row=current_row, column=ci, value=val)
        _style_section_header(ws, current_row, "FEF3C7")
        current_row += 1

        for i, match in enumerate(round_matches):
            row_data = [""] * _NCOLS
            row_data[_COL_ORA]      = _format_time(match.scheduled_at, tournament_timezone)
            row_data[_COL_CAMPO]    = _field_label(match)
            row_data[_COL_NOM_HOME] = match.home_label or ""
            row_data[_COL_VS]       = "vs"
            row_data[_COL_NOM_AWAY] = match.away_label or ""
            row_data[_COL_PT_H]   = match.home_score if match.home_score is not None else ""
            row_data[_COL_PT_A]   = match.away_score if match.away_score is not None else ""
            row_data[_COL_ARBITRO]  = match.referee or ""
            for ci, val in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=ci, value=val)
            _style_match_row(ws, current_row, i % 2 == 1)
            current_row += 1

        # Spacer between rounds
        ws.row_dimensions[current_row].height = 8
        current_row += 1


# ── Block writers (offset-aware, for single-sheet layout) ─────────────────────

def _write_group_block(
    ws,
    start_row: int,
    age_group_label: str,
    group: "ProgramGroupResponse",
    group_num: int,
    phase_name: str,
    tournament_timezone: str,
    image_cache: dict,
) -> int:
    """Write one group (team list + matches) starting at *start_row*. Returns next free row."""
    from openpyxl.styles import Font, Alignment

    team_code: dict[str, tuple[str, str]] = {}
    for pos, team in enumerate(group.teams):
        code = f"{chr(65 + pos)}{group_num}"
        if team.tournament_team_id:
            team_code[team.tournament_team_id] = (code, team.label)

    r = start_row

    # Group name banner
    ws.row_dimensions[r].height = 20
    ws.cell(row=r, column=1, value=group.name)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_NCOLS)
    _style_group_name_row(ws, r)
    r += 1

    # Team-roster header
    ws.row_dimensions[r].height = 16
    for ci, val in enumerate(["COD.", "SQUADRA"] + [""] * (_NCOLS - 2), start=1):
        ws.cell(row=r, column=ci, value=val)
    _style_section_header(ws, r, "DBEAFE")
    r += 1

    # Team rows
    for pos, team in enumerate(group.teams):
        code = f"{chr(65 + pos)}{group_num}"
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=team.label)
        _style_match_row(ws, r, pos % 2 == 1, row_height=26)
        logo_bytes = _fetch_image_bytes(getattr(team, "team_logo_url", None), image_cache)
        _add_image(ws, logo_bytes, f"C{r}", 20, 20)
        r += 1

    # Spacer
    ws.row_dimensions[r].height = 6
    r += 1

    # Match header
    header = [""] * _NCOLS
    header[_COL_ORA]      = "ORA"
    header[_COL_CAMPO]    = "CAMPO"
    header[_COL_NOM_HOME] = "SQUADRA CASA"
    header[_COL_VS]       = "vs"
    header[_COL_NOM_AWAY] = "SQUADRA OSPITE"
    header[_COL_PT_H]     = "PT"
    header[_COL_PT_A]     = "PT"
    header[_COL_ARBITRO]  = "ARBITRO"
    for ci, val in enumerate(header, start=1):
        ws.cell(row=r, column=ci, value=val)
    _style_section_header(ws, r, "DBEAFE")
    r += 1

    # Match rows
    for i, match in enumerate(_sort_matches(group.matches)):
        h_code, h_name = team_code.get(match.home_team_id or "", ("", match.home_label or ""))
        a_code, a_name = team_code.get(match.away_team_id or "", ("", match.away_label or ""))
        row_data = [""] * _NCOLS
        row_data[_COL_ORA]      = _format_time(match.scheduled_at, tournament_timezone)
        row_data[_COL_CAMPO]    = _field_label(match)
        row_data[_COL_COD_HOME] = h_code
        row_data[_COL_NOM_HOME] = h_name
        row_data[_COL_VS]       = "vs"
        row_data[_COL_COD_AWAY] = a_code
        row_data[_COL_NOM_AWAY] = a_name
        row_data[_COL_PT_H]     = match.home_score if match.home_score is not None else ""
        row_data[_COL_PT_A]     = match.away_score if match.away_score is not None else ""
        row_data[_COL_ARBITRO]  = match.referee or ""
        for ci, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=ci, value=val)
        _style_match_row(ws, r, i % 2 == 1)
        if match.home_logo_url:
            logo = _fetch_image_bytes(match.home_logo_url, image_cache)
            _add_image(ws, logo, f"C{r}", 14, 14)
        r += 1

    # Trailing spacer
    ws.row_dimensions[r].height = 10
    r += 1
    return r


def _write_knockout_block(
    ws,
    start_row: int,
    phase: "ProgramPhaseResponse",
    tournament_timezone: str,
) -> int:
    """Write knockout rounds starting at *start_row*. Returns next free row."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    r = start_row

    # Phase name banner (amber-dark, same style as round header in _write_knockout_sheet)
    ws.row_dimensions[r].height = 20
    amber_fill = PatternFill("solid", fgColor="92400E")
    thin = Side(style="thin", color="B45309")
    for col in range(1, _NCOLS + 1):
        c = ws.cell(row=r, column=col)
        c.fill = amber_fill
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = ws.cell(row=r, column=1, value=phase.name.upper())
    cell.font = Font(bold=True, size=12, color="FEF3C7", name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_NCOLS)
    r += 1

    matches_by_round: dict[str, list] = {}
    for m in _sort_matches(phase.knockout_matches):
        matches_by_round.setdefault(m.bracket_round or phase.name, []).append(m)

    for round_name, round_matches in matches_by_round.items():
        # Round sub-header
        ws.row_dimensions[r].height = 16
        light_amber = PatternFill("solid", fgColor="B45309")
        for col in range(1, _NCOLS + 1):
            c = ws.cell(row=r, column=col)
            c.fill = light_amber
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        rnd_cell = ws.cell(
            row=r,
            column=1,
            value=f"{round_name.upper()}{_round_duration_suffix(round_matches, phase.match_duration_minutes)}",
        )
        rnd_cell.font = Font(bold=True, size=10, color="FEF3C7", name="Calibri")
        rnd_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_NCOLS)
        r += 1

        # Column header
        header = [""] * _NCOLS
        header[_COL_ORA]      = "ORA"
        header[_COL_CAMPO]    = "CAMPO"
        header[_COL_NOM_HOME] = "SQUADRA CASA"
        header[_COL_VS]       = "vs"
        header[_COL_NOM_AWAY] = "SQUADRA OSPITE"
        header[_COL_PT_H]     = "PT"
        header[_COL_PT_A]     = "PT"
        header[_COL_ARBITRO]  = "ARBITRO"
        for ci, val in enumerate(header, start=1):
            ws.cell(row=r, column=ci, value=val)
        _style_section_header(ws, r, "FEF3C7")
        r += 1

        for i, match in enumerate(round_matches):
            row_data = [""] * _NCOLS
            row_data[_COL_ORA]      = _format_time(match.scheduled_at, tournament_timezone)
            row_data[_COL_CAMPO]    = _field_label(match)
            row_data[_COL_NOM_HOME] = match.home_label or ""
            row_data[_COL_VS]       = "vs"
            row_data[_COL_NOM_AWAY] = match.away_label or ""
            row_data[_COL_PT_H]     = match.home_score if match.home_score is not None else ""
            row_data[_COL_PT_A]     = match.away_score if match.away_score is not None else ""
            row_data[_COL_ARBITRO]  = match.referee or ""
            for ci, val in enumerate(row_data, start=1):
                ws.cell(row=r, column=ci, value=val)
            _style_match_row(ws, r, i % 2 == 1)
            r += 1

        ws.row_dimensions[r].height = 8
        r += 1

    return r


# ── Public entry point ─────────────────────────────────────────────────────────

def build_age_group_program_excel(
    tournament_name: str,
    program: AgeGroupProgramResponse,
    tournament_timezone: str = "Europe/Rome",
    organization_logo_url: str | None = None,
    tournament_logo_url: str | None = None,
) -> tuple[bytes, str]:
    from openpyxl import Workbook

    age_group_label = program.display_name or program.age_group
    image_cache: dict = {}

    org_logo        = _fetch_image_bytes(organization_logo_url, image_cache)
    tournament_logo = _fetch_image_bytes(tournament_logo_url, image_cache)

    wb = Workbook()
    wb.remove(wb.active)

    for day in program.days:
        for phase in day.phases:
            dur_suffix = f"  ·  {phase.match_duration_minutes} min/partita" if phase.match_duration_minutes else ""
            if phase.groups:
                for gi, group in enumerate(phase.groups):
                    sheet_name = _safe_sheet_name(
                        f"G{gi + 1} - {phase.name}" if len(phase.groups) > 1 else phase.name
                    )
                    ws = wb.create_sheet(title=sheet_name)
                    _write_group_sheet(
                        ws, tournament_name, age_group_label, group,
                        gi + 1, f"{phase.name}{dur_suffix}", tournament_timezone,
                        org_logo, tournament_logo, image_cache,
                        match_duration_minutes=phase.match_duration_minutes,
                    )
            if phase.knockout_matches:
                ws = wb.create_sheet(title=_safe_sheet_name(phase.name))
                _write_knockout_sheet(
                    ws, tournament_name, age_group_label, phase,
                    tournament_timezone, org_logo, tournament_logo,
                )

    if not wb.sheetnames:
        wb.create_sheet("Programma")

    buffer = BytesIO()
    wb.save(buffer)
    filename = f"programma-{_safe_filename(tournament_name)}-{_safe_filename(age_group_label)}.xlsx"
    return buffer.getvalue(), filename


# ── Full tournament Excel (all age groups, one workbook) ──────────────────────

def build_full_tournament_excel(
    tournament_name: str,
    programs: list[AgeGroupProgramResponse],
    tournament_timezone: str = "Europe/Rome",
    organization_logo_url: str | None = None,
    tournament_logo_url: str | None = None,
) -> tuple[bytes, str]:
    from openpyxl import Workbook

    image_cache: dict = {}
    org_logo        = _fetch_image_bytes(organization_logo_url, image_cache)
    tournament_logo = _fetch_image_bytes(tournament_logo_url, image_cache)

    wb = Workbook()
    wb.remove(wb.active)

    used_sheet_names: set[str] = set()

    def _unique_sheet_name(name: str) -> str:
        if name not in used_sheet_names:
            used_sheet_names.add(name)
            return name
        i = 2
        while True:
            candidate = _safe_sheet_name(f"{name[:27]} {i}")
            if candidate not in used_sheet_names:
                used_sheet_names.add(candidate)
                return candidate
            i += 1

    for program in programs:
        age_group_label = program.display_name or program.age_group
        prefix = age_group_label[:10]
        for day in program.days:
            for phase in day.phases:
                dur_suffix = f"  ·  {phase.match_duration_minutes} min/partita" if phase.match_duration_minutes else ""
                if phase.groups:
                    for gi, group in enumerate(phase.groups):
                        raw = (
                            f"{prefix} · G{gi + 1}" if len(phase.groups) > 1
                            else f"{prefix} · {phase.name}"
                        )
                        sheet_name = _unique_sheet_name(_safe_sheet_name(raw))
                        ws = wb.create_sheet(title=sheet_name)
                        _write_group_sheet(
                            ws, tournament_name, age_group_label, group,
                            gi + 1, f"{phase.name}{dur_suffix}", tournament_timezone,
                            org_logo, tournament_logo, image_cache,
                            match_duration_minutes=phase.match_duration_minutes,
                        )
                if phase.knockout_matches:
                    raw = f"{prefix} · {phase.name}"
                    sheet_name = _unique_sheet_name(_safe_sheet_name(raw))
                    ws = wb.create_sheet(title=sheet_name)
                    _write_knockout_sheet(
                        ws, tournament_name, age_group_label, phase,
                        tournament_timezone, org_logo, tournament_logo,
                    )

    if not wb.sheetnames:
        wb.create_sheet("Programma")

    buffer = BytesIO()
    wb.save(buffer)
    filename = f"programma-completo-{_safe_filename(tournament_name)}.xlsx"
    return buffer.getvalue(), filename


# ── Campo calendar Excel (all matches grouped by field) ───────────────────────

_CAMPO_NCOLS = 8

def _write_campo_sheet(
    ws,
    tournament_name: str,
    field_label_str: str,
    entries: list[dict],
    tournament_timezone: str,
    org_logo: bytes | None,
    tournament_logo: bytes | None,
) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    col_widths = {"A": 8, "B": 30, "C": 8, "D": 26, "E": 4, "F": 26, "G": 20, "H": 22}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Row 1: header ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 52
    navy_fill = PatternFill("solid", fgColor="1E3A5F")
    thin = Side(style="thin", color="2D5A8E")
    for col in range(1, _CAMPO_NCOLS + 1):
        c = ws.cell(row=1, column=col)
        c.fill = navy_fill
        c.border = Border(bottom=thin)

    title_cell = ws.cell(row=1, column=2, value=tournament_name.upper())
    title_cell.font = Font(bold=True, size=18, color="FFFFFF", name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
    _add_image(ws, org_logo, "A1", 42, 42)
    _add_image(ws, tournament_logo, "H1", 42, 42)

    # ── Row 2: field name ─────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 22
    field_fill = PatternFill("solid", fgColor="1D4ED8")
    for col in range(1, _CAMPO_NCOLS + 1):
        ws.cell(row=2, column=col).fill = field_fill
    fc = ws.cell(row=2, column=1, value=field_label_str.upper())
    fc.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    fc.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_CAMPO_NCOLS)

    # ── Row 3: accent bar ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 5
    accent = PatternFill("solid", fgColor="F59E0B")
    for col in range(1, _CAMPO_NCOLS + 1):
        ws.cell(row=3, column=col).fill = accent

    # ── Row 4: column headers ─────────────────────────────────────────────────
    headers = ["ORA", "CATEGORIA / GIRONE", "CAMPO", "SQUADRA CASA", "vs", "SQUADRA OSPITE", "ARBITRO", "PT/PT"]
    for ci, hdr in enumerate(headers, start=1):
        ws.cell(row=4, column=ci, value=hdr)
    from openpyxl.styles import Font as Fnt, PatternFill as PF, Alignment as Al, Border as Bd, Side as Sd
    header_fill = PF("solid", fgColor="DBEAFE")
    bthin = Sd(style="thin", color="93C5FD")
    bborder = Bd(left=bthin, right=bthin, top=bthin, bottom=Sd(style="medium", color="3B82F6"))
    ws.row_dimensions[4].height = 18
    for col in range(1, _CAMPO_NCOLS + 1):
        c = ws.cell(row=4, column=col)
        c.fill = header_fill
        c.border = bborder
        c.font = Fnt(bold=True, size=9, color="1E3A5F", name="Calibri")
        c.alignment = Al(horizontal="center", vertical="center")

    # ── Match rows ────────────────────────────────────────────────────────────
    thin_s = Sd(style="thin", color="CBD5E1")
    row_border = Bd(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    for i, entry in enumerate(sorted(entries, key=lambda e: e.get("time_sort", ""))):
        r = 5 + i
        ws.row_dimensions[r].height = 18
        alt_fill = PF("solid", fgColor="F0F4FF" if i % 2 == 0 else "FFFFFF")
        score = ""
        if entry.get("home_score") is not None and entry.get("away_score") is not None:
            score = f"{entry['home_score']} – {entry['away_score']}"
        row_vals = [
            entry.get("time", ""),
            entry.get("context", ""),
            entry.get("field", ""),
            entry.get("home", ""),
            "vs",
            entry.get("away", ""),
            entry.get("referee", ""),
            score,
        ]
        for ci, val in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = alt_fill
            c.border = row_border
            c.font = Fnt(size=9, name="Calibri")
            c.alignment = Al(vertical="center")
        ws.cell(row=r, column=1).alignment = Al(horizontal="center", vertical="center")
        ws.cell(row=r, column=5).alignment = Al(horizontal="center", vertical="center")
        ws.cell(row=r, column=8).alignment = Al(horizontal="center", vertical="center")


def build_tournament_campo_calendar_excel(
    tournament_name: str,
    programs: list[AgeGroupProgramResponse],
    tournament_timezone: str = "Europe/Rome",
    organization_logo_url: str | None = None,
    tournament_logo_url: str | None = None,
) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from zoneinfo import ZoneInfo

    image_cache: dict = {}
    org_logo        = _fetch_image_bytes(organization_logo_url, image_cache)
    tournament_logo = _fetch_image_bytes(tournament_logo_url, image_cache)

    try:
        tz = ZoneInfo(tournament_timezone)
    except Exception:
        tz = ZoneInfo("Europe/Rome")

    # Collect all matches keyed by field
    field_entries: dict[str, list[dict]] = {}
    field_labels: dict[str, str] = {}

    for program in programs:
        ag_label = program.display_name or program.age_group
        for day in program.days:
            for phase in day.phases:
                all_matches: list = []
                for group in phase.groups:
                    for m in group.matches:
                        all_matches.append((m, f"{ag_label} · {group.name}"))
                for m in phase.knockout_matches:
                    rnd = m.bracket_round or phase.name
                    all_matches.append((m, f"{ag_label} · {rnd}"))

                for match, context in all_matches:
                    if not match.field_name:
                        continue
                    key = f"{match.field_name}::{match.field_number if match.field_number is not None else ''}"
                    if key not in field_entries:
                        field_entries[key] = []
                        if match.field_number is not None:
                            field_labels[key] = f"{match.field_name} · Campo {match.field_number}"
                        else:
                            field_labels[key] = match.field_name

                    time_str = ""
                    time_sort = ""
                    if match.scheduled_at:
                        try:
                            if getattr(match.scheduled_at, "tzinfo", None) is None:
                                time_str = match.scheduled_at.strftime("%H:%M")
                                time_sort = match.scheduled_at.isoformat()
                            else:
                                local = match.scheduled_at.astimezone(tz)
                                time_str = local.strftime("%H:%M")
                                time_sort = local.isoformat()
                        except Exception:
                            pass

                    field_entries[key].append({
                        "time": time_str,
                        "time_sort": time_sort,
                        "context": context,
                        "field": f"Campo {match.field_number}" if match.field_number is not None else "",
                        "home": match.home_label or "",
                        "away": match.away_label or "",
                        "referee": match.referee or "",
                        "home_score": match.home_score,
                        "away_score": match.away_score,
                    })

    wb = Workbook()
    wb.remove(wb.active)

    for key in sorted(field_entries.keys()):
        sheet_name = _safe_sheet_name(field_labels.get(key, key))
        ws = wb.create_sheet(title=sheet_name)
        _write_campo_sheet(
            ws, tournament_name, field_labels.get(key, key),
            field_entries[key], tournament_timezone, org_logo, tournament_logo,
        )

    if not wb.sheetnames:
        wb.create_sheet("Calendario")

    buffer = BytesIO()
    wb.save(buffer)
    filename = f"calendario-impianti-{_safe_filename(tournament_name)}.xlsx"
    return buffer.getvalue(), filename
